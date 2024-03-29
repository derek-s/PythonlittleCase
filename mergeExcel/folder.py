#!/usr/bin/python
# -*- coding: utf-8 -*-
# @Time    : 2018/11/29 11:06
# @Author  : Derek.S
# @Site    : 
# @File    : folder.py

import glob
import openpyxl


def openExcel(filename):
    excel = openpyxl.load_workbook(filename)
    return excel


def getRow(workbook, sheet, listIndex):

    oldSheet = workbook[str(sheet)]
    contents = []

    # 循环得到每一行的值
    for value in oldSheet.values:
        contents.append(list(value))

    # 如果不是文件列表中的第1个文件，删除列名行
    if listIndex != 0:
        del contents[0]

    return contents


def writeNewFile(workbook, sheet, contents):
    try:
        ws = workbook[str(sheet)]
    except:
        ws = workbook.create_sheet(title=sheet)
    for eachRow in contents:
        ws.append(eachRow)




if __name__ == "__main__":

    fileNameList = []
    outputFileName = "hb.xlsx"


    for filename in glob.glob(r'*.xlsx'):
        fileNameList.append(filename)

    # 创建空Excel文件用于合并

    megerExcel = openpyxl.Workbook()

    # 读取待合并文件并合并到新文件

    for eachFile in fileNameList:
        FNamelistIndex = fileNameList.index(eachFile)
        excelFile = openExcel(eachFile)
        for eachSheet in excelFile.sheetnames:
            oldExcel = getRow(excelFile, eachSheet, FNamelistIndex)
            writeNewFile(megerExcel, eachSheet, oldExcel)

    defaultSheet = megerExcel["Sheet"]
    megerExcel.remove_sheet(defaultSheet)
    megerExcel.save(outputFileName)
